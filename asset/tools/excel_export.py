from django.test import TestCase

# Create your tests here.
from openpyxl import load_workbook
import re
import os

def xslx_data():

    # 定义上传路径
    #base_dir = '/home/changhao/PycharmProjects/northcorezh/northcorePro_v1.0/test/northcorecs/static/excel_download'
    base_dir = 'E:\\code\\luffy\\northcorezh\\northcorecs\\static\\excel_download\\'
    a = os.listdir(base_dir)

    # 接收excel文件名
    xslx_name = ''

    for i in a:
        if re.findall("xlsx", i):
            xslx_name += i

    # 加载excel
    wb = load_workbook(base_dir + xslx_name)

    # 激活sheet页
    sheet = wb.active

    # print(wb.get_sheet_names())  工作表1
    # wb.get_sheet_by_name('工作表1')

    # print(sheet.max_column)  #查看最大列
    # print(sheet.max_row)     #查看最大行

    # 将excel数据转换字典格式存放
    data = {}
    zy_list = []

    # 获取最大行
    row_line = sheet.max_row + 1

    # 遍历excel所有数据,存放到字典
    for i in range(2, row_line):

        for row in sheet.iter_rows('A{0}:K{1}'.format(i, i)):

            for call in row:
                zy_list.append(call.value)

            data[i] = str(zy_list)
            zy_list.clear()

        i += 1

    # 查看字典
    # for dict in data:
    #     print(data[dict])

    # 删除上传后的excel
    os.remove(base_dir+xslx_name)

    return data


def export_data():

    #1. 将当前数据库的设备信息  -->  字典或列表格式
    #2. 将字典或列表 写入到excel

    pass


